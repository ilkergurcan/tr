“””
Excel preprocessing & chunking pipeline for RAG (v2 — auto-detection).

Handles varied Excel layouts without manual row/col configuration:

- Auto-detects header row, data start row, and footnote sections
- Handles horizontal merged cells (section headers)
- Handles vertical merged cells (subtopic labels via forward-fill)
- Resolves footnote/citation markers inline or as metadata
- Works on both messy (merged, footnoted) and clean Excel files

Dependencies: openpyxl, pandas
“””

import re
from dataclasses import dataclass, field
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# —————————————————————————

# Data classes

# —————————————————————————

@dataclass
class MergeInfo:
value: str | None
min_row: int
max_row: int
min_col: int
max_col: int

```
@property
def col_span(self) -> int:
    return self.max_col - self.min_col + 1

@property
def row_span(self) -> int:
    return self.max_row - self.min_row + 1

@property
def is_horizontal_header(self) -> bool:
    return self.row_span == 1 and self.col_span > 1

@property
def is_vertical_label(self) -> bool:
    return self.col_span == 1 and self.row_span > 1
```

@dataclass
class Chunk:
text: str
metadata: dict = field(default_factory=dict)

# —————————————————————————

# Footnote patterns

# —————————————————————————

FOOTNOTE_MARKER_RE = re.compile(
r’((\d+))’
r’|((*+))’
r’|(([!]+))’
r’|[(\d+)]’
r’|([¹²³⁴⁵⁶⁷⁸⁹⁰†‡§])’
)

FOOTNOTE_DEF_RE = re.compile(
r’^\s*((\d+)|(*+)|([!]+)|[\d+]|[¹²³⁴⁵⁶⁷⁸⁹⁰†‡§])\s*(.+)’
)

FOOTNOTE_SECTION_KEYWORDS = re.compile(
r’^\s*(not(?:e|es)|source|sources|reference|references|legend|footnot|açıklama|kaynak)’,
re.IGNORECASE,
)

# —————————————————————————

# Step 1: Catalog & unmerge

# —————————————————————————

def catalog_merges(ws) -> list[MergeInfo]:
catalog = []
for mr in list(ws.merged_cells.ranges):
min_col, min_row, max_col, max_row = range_boundaries(str(mr))
catalog.append(MergeInfo(
value=ws.cell(row=min_row, column=min_col).value,
min_row=min_row, max_row=max_row,
min_col=min_col, max_col=max_col,
))
return catalog

def unmerge_and_fill(ws, catalog: list[MergeInfo]) -> None:
for mr in list(ws.merged_cells.ranges):
min_col, min_row, max_col, max_row = range_boundaries(str(mr))
value = ws.cell(row=min_row, column=min_col).value
ws.unmerge_cells(str(mr))
for row in ws.iter_rows(min_col=min_col, min_row=min_row,
max_col=max_col, max_row=max_row):
for cell in row:
cell.value = value

# —————————————————————————

# Step 2: Auto-detect header row and data start

# —————————————————————————

def _row_values(ws, row_idx: int, min_col: int, max_col: int) -> list:
return [ws.cell(row=row_idx, column=c).value for c in range(min_col, max_col + 1)]

def _non_none_count(values: list) -> int:
return sum(1 for v in values if v is not None)

def _is_likely_header(values: list) -> bool:
“”“A header row is mostly short strings (not numbers, not long text).”””
non_none = [v for v in values if v is not None]
if len(non_none) < 2:
return False
string_count = sum(
1 for v in non_none
if isinstance(v, str) and len(v.strip()) < 80
)
# Most cells should be short strings
return string_count / len(non_none) >= 0.6

def _is_section_header_row(row_idx: int, catalog: list[MergeInfo], data_width: int) -> bool:
“”“Check if this row is a wide horizontal merge (section header).”””
for m in catalog:
if m.min_row == row_idx and m.is_horizontal_header and m.col_span >= data_width * 0.4:
return True
return False

def _is_number_or_data(v) -> bool:
“”“Check if a value looks like actual data (number, date, long text).”””
if v is None:
return False
if isinstance(v, (int, float)):
return True
if isinstance(v, str):
# Looks numeric
cleaned = v.strip().replace(”,”, “”).replace(”.”, “”).replace(”%”, “”).replace(”$”, “”).replace(“€”, “”).replace(“₺”, “”)
if cleaned.isdigit():
return True
# Long text = likely data, not header
if len(v.strip()) > 80:
return True
return False

def detect_header_and_data_start(
ws, catalog: list[MergeInfo]
) -> tuple[int | None, int]:
“””
Auto-detect the column header row and first data row.

```
Strategy:
- Skip section-header rows (wide horizontal merges)
- Skip empty rows
- First row that looks like a header (mostly short strings, 2+ filled cells) = header
- First row after header that contains data = data start
- If no header found, assume row 1 is header

Returns:
    (header_row, data_start_row)  — header_row can be None if no clear header
"""
min_col, max_col = ws.min_column, ws.max_column
data_width = max_col - min_col + 1
max_row = ws.max_row

if max_row is None or max_row == 0:
    return None, 1

header_row = None

for row_idx in range(ws.min_row, min(max_row + 1, 50)):  # scan first 50 rows max
    vals = _row_values(ws, row_idx, min_col, max_col)

    # Skip fully empty rows
    if _non_none_count(vals) == 0:
        continue

    # Skip section header rows (wide merges)
    if _is_section_header_row(row_idx, catalog, data_width):
        continue

    # Check if this looks like a column header row
    if _is_likely_header(vals):
        # Verify the NEXT non-empty row has actual data (numbers, longer text, etc.)
        # This avoids misidentifying a data row as a header
        for next_row in range(row_idx + 1, min(max_row + 1, row_idx + 10)):
            next_vals = _row_values(ws, next_row, min_col, max_col)
            if _non_none_count(next_vals) == 0:
                continue
            # If next row has numbers/data, current row is likely the header
            if any(_is_number_or_data(v) for v in next_vals):
                header_row = row_idx
                break
            # If next row also looks like headers, skip (multi-row header situation)
            if _is_likely_header(next_vals):
                continue
            else:
                header_row = row_idx
                break

        if header_row:
            break

# Find data start: first non-empty row after header
data_start = (header_row or ws.min_row) + 1
for row_idx in range(data_start, min(max_row + 1, data_start + 20)):
    vals = _row_values(ws, row_idx, min_col, max_col)
    if _non_none_count(vals) > 0:
        # Skip if it's another section header
        if _is_section_header_row(row_idx, catalog, data_width):
            continue
        data_start = row_idx
        break

return header_row, data_start
```

# —————————————————————————

# Step 3: Hierarchy from horizontal headers

# —————————————————————————

def classify_horizontal_headers(
catalog: list[MergeInfo],
data_width: int,
width_threshold: float = 0.4,
) -> list[MergeInfo]:
headers = [
m for m in catalog
if m.is_horizontal_header and m.col_span >= data_width * width_threshold
]
return sorted(headers, key=lambda m: -m.col_span)

def get_hierarchy_for_row(row_num: int, headers: list[MergeInfo]) -> list[str]:
“””
For same-width headers (same hierarchy level), only the closest one
above the current row applies. E.g., “Financial Benefits” replaces
“Health Benefits” once you pass that section boundary.
“””
# Group by col_span (proxy for hierarchy level), pick closest above
best_by_level: dict[int, MergeInfo] = {}
for h in headers:
if h.value and h.min_row < row_num:
level = h.col_span
if level not in best_by_level or h.min_row > best_by_level[level].min_row:
best_by_level[level] = h

```
# Sort by col_span descending (broadest = top-level)
sorted_levels = sorted(best_by_level.items(), key=lambda x: -x[0])
return [str(m.value).strip() for _, m in sorted_levels]
```

# —————————————————————————

# Step 4: Footnote detection & resolution

# —————————————————————————

def find_footnote_section_start(ws, data_start_row: int) -> int | None:
“””
Scan bottom-up to find where footnotes begin.
Only looks BELOW the data start row.
Added safeguard: won’t mark footnotes unless they actually match definition patterns.
“””
max_row = ws.max_row
max_col = ws.max_column

```
if max_row is None:
    return None

# Strategy 1: Look for keyword markers ("Notes:", "Source:", etc.)
for row_idx in range(max_row, data_start_row, -1):
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val and FOOTNOTE_SECTION_KEYWORDS.match(str(val)):
            return row_idx

# Strategy 2: Scan bottom-up for consecutive footnote-definition rows
footnote_top = None
for row_idx in range(max_row, data_start_row, -1):
    row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
    combined = " ".join(str(v) for v in row_vals if v).strip()

    if not combined:
        # Empty row — if we already found footnotes below, this gap confirms the boundary
        if footnote_top is not None:
            break
        continue

    if FOOTNOTE_DEF_RE.match(combined):
        footnote_top = row_idx
    else:
        # Non-footnote row — stop scanning
        break

return footnote_top
```

def build_footnote_registry(ws, footnote_start_row: int) -> dict[str, str]:
registry = {}
max_row = ws.max_row
max_col = ws.max_column

```
current_marker = None
current_text = []

for row_idx in range(footnote_start_row, max_row + 1):
    row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
    line = " ".join(str(v) for v in row_vals if v).strip()

    if not line:
        continue

    # Skip the keyword row itself ("Notes:", "Source:", etc.)
    if FOOTNOTE_SECTION_KEYWORDS.match(line) and not FOOTNOTE_DEF_RE.match(line):
        continue

    match = FOOTNOTE_DEF_RE.match(line)
    if match:
        if current_marker:
            registry[current_marker] = " ".join(current_text).strip()
        current_marker = match.group(1)
        current_text = [match.group(2)]
    elif current_marker:
        current_text.append(line)

if current_marker:
    registry[current_marker] = " ".join(current_text).strip()

return registry
```

def resolve_footnotes_in_text(
text: str,
registry: dict[str, str],
inline: bool = True,
) -> tuple[str, dict[str, str]]:
resolved = text
found_footnotes = {}
for marker, explanation in registry.items():
if marker in text:
found_footnotes[marker] = explanation
if inline:
resolved = resolved.replace(marker, f” [Note: {explanation}]”)
return resolved, found_footnotes

# —————————————————————————

# Step 5: Chunk the sheet

# —————————————————————————

def get_column_headers(ws, header_row: int | None, min_col: int, max_col: int) -> dict[int, str]:
if header_row is None:
return {col: f”Column_{col}” for col in range(min_col, max_col + 1)}
headers = {}
for col in range(min_col, max_col + 1):
val = ws.cell(row=header_row, column=col).value
headers[col] = str(val).strip() if val else f”Column_{col}”
return headers

def chunk_sheet(
ws,
catalog: list[MergeInfo],
footnote_registry: dict[str, str],
header_row: int | None,
data_start_row: int,
footnote_start_row: int | None,
inline_footnotes: bool,
sheet_name: str,
) -> list[Chunk]:
min_col, max_col = ws.min_column, ws.max_column
data_width = max_col - min_col + 1
max_row = ws.max_row

```
if max_row is None:
    return []

col_headers = get_column_headers(ws, header_row, min_col, max_col)
h_headers = classify_horizontal_headers(catalog, data_width)
last_data_row = (footnote_start_row - 1) if footnote_start_row else max_row

# Pre-compute secondary header rows: the FIRST header-like row directly after each section header
secondary_header_rows: set[int] = set()
section_header_rows = {
    m.min_row for m in catalog
    if m.is_horizontal_header and m.col_span >= data_width * 0.4
}
for sh_row in section_header_rows:
    # Check the next few rows for a header-like row
    for candidate in range(sh_row + 1, min(sh_row + 4, (last_data_row or max_row) + 1)):
        candidate_vals = [
            ws.cell(row=candidate, column=c).value
            for c in range(min_col, max_col + 1)
        ]
        non_none = [v for v in candidate_vals if v is not None]
        if not non_none:
            continue  # skip empty rows between section header and sub-header
        if _is_likely_header(non_none) and not any(_is_number_or_data(v) for v in non_none):
            secondary_header_rows.add(candidate)
        break  # only check until first non-empty row

chunks = []

for row_idx in range(data_start_row, last_data_row + 1):
    row_vals = {
        col: ws.cell(row=row_idx, column=col).value
        for col in range(min_col, max_col + 1)
    }

    if all(v is None for v in row_vals.values()):
        continue

    # Skip rows that are themselves section headers (wide merges)
    if _is_section_header_row(row_idx, catalog, data_width):
        continue

    # Skip pre-identified secondary header rows and update col_headers
    if row_idx in secondary_header_rows:
        for col in range(min_col, max_col + 1):
            val = row_vals.get(col)
            if val:
                col_headers[col] = str(val).strip()
        continue

    hierarchy = get_hierarchy_for_row(row_idx, h_headers)

    parts = []
    row_footnotes = {}

    for col in range(min_col, max_col + 1):
        val = row_vals.get(col)
        if val is None:
            continue

        col_name = col_headers.get(col, f"Column_{col}")
        cell_text = str(val).strip()

        if footnote_registry:
            cell_text, cell_fn = resolve_footnotes_in_text(
                cell_text, footnote_registry, inline=inline_footnotes
            )
            row_footnotes.update(cell_fn)

        parts.append(f"{col_name}: {cell_text}")

    if not parts:
        continue

    prefix = " > ".join(hierarchy)
    row_text = " | ".join(parts)
    chunk_text = f"[{prefix}] {row_text}" if prefix else row_text

    metadata = {
        "source_type": "excel",
        "sheet_name": sheet_name,
        "row": row_idx,
        "hierarchy": hierarchy,
    }
    if row_footnotes and not inline_footnotes:
        metadata["footnotes"] = row_footnotes

    chunks.append(Chunk(text=chunk_text, metadata=metadata))

return chunks
```

# —————————————————————————

# Main pipeline

# —————————————————————————

def process_excel(
filepath: str,
inline_footnotes: bool = True,
sheets: list[str] | None = None,
) -> list[Chunk]:
“””
Full auto-detection pipeline.
No manual header_row / data_start_row needed.

```
Args:
    filepath: Path to .xlsx file
    inline_footnotes: Inline footnote text into chunks (recommended True)
    sheets: Sheet names to process. None = all sheets.

Returns:
    List of Chunk objects ready for embedding.
"""
wb = load_workbook(filepath, data_only=True)
sheet_names = sheets or wb.sheetnames
all_chunks = []

for sheet_name in sheet_names:
    if sheet_name not in wb.sheetnames:
        print(f"[WARN] Sheet '{sheet_name}' not found, skipping.")
        continue

    ws = wb[sheet_name]

    # Skip empty sheets
    if ws.max_row is None or ws.max_row == 0:
        print(f"[SKIP] Sheet '{sheet_name}' is empty.")
        continue

    # Step 1: Catalog merges before unmerging
    catalog = catalog_merges(ws)

    # Step 2: Unmerge and propagate values
    unmerge_and_fill(ws, catalog)

    # Step 3: Auto-detect layout
    header_row, data_start_row = detect_header_and_data_start(ws, catalog)

    # Step 4: Detect footnotes (only below data start)
    footnote_start = find_footnote_section_start(ws, data_start_row)
    footnote_registry = {}
    if footnote_start:
        footnote_registry = build_footnote_registry(ws, footnote_start)

    # Debug logging
    print(
        f"[INFO] Sheet '{sheet_name}': "
        f"header_row={header_row}, data_start={data_start_row}, "
        f"footnote_start={footnote_start}, merges={len(catalog)}, "
        f"footnotes={len(footnote_registry)}"
    )

    # Step 5: Chunk
    chunks = chunk_sheet(
        ws=ws,
        catalog=catalog,
        footnote_registry=footnote_registry,
        header_row=header_row,
        data_start_row=data_start_row,
        footnote_start_row=footnote_start,
        inline_footnotes=inline_footnotes,
        sheet_name=sheet_name,
    )
    all_chunks.extend(chunks)

if not all_chunks:
    print("[WARN] No chunks produced. Check the [INFO] logs above for detection results.")

return all_chunks
```

# —————————————————————————

# Debug utility — run this on your file to see what the detector finds

# —————————————————————————

def debug_sheet(filepath: str, sheet_name: str | None = None):
“”“Print detected layout for each sheet — use this to diagnose issues.”””
wb = load_workbook(filepath, data_only=True)
sheets = [sheet_name] if sheet_name else wb.sheetnames

```
for sname in sheets:
    ws = wb[sname]
    print(f"\n{'='*60}")
    print(f"Sheet: {sname}")
    print(f"Dimensions: rows={ws.min_row}-{ws.max_row}, cols={ws.min_column}-{ws.max_column}")

    catalog = catalog_merges(ws)
    print(f"Merged ranges: {len(catalog)}")
    for m in catalog[:10]:  # show first 10
        mtype = "H-HEADER" if m.is_horizontal_header else ("V-LABEL" if m.is_vertical_label else "BLOCK")
        print(f"  [{mtype}] row {m.min_row}-{m.max_row}, col {m.min_col}-{m.max_col} = '{m.value}'")

    unmerge_and_fill(ws, catalog)
    header_row, data_start = detect_header_and_data_start(ws, catalog)

    print(f"\nDetected header row: {header_row}")
    if header_row:
        vals = _row_values(ws, header_row, ws.min_column, ws.max_column)
        print(f"  Header values: {[v for v in vals if v]}")

    print(f"Detected data start: {data_start}")
    if data_start and data_start <= ws.max_row:
        vals = _row_values(ws, data_start, ws.min_column, ws.max_column)
        print(f"  First data row: {[v for v in vals if v]}")

    fn_start = find_footnote_section_start(ws, data_start)
    print(f"Detected footnote start: {fn_start}")
    if fn_start:
        registry = build_footnote_registry(ws, fn_start)
        print(f"  Footnotes found: {len(registry)}")
        for marker, text in list(registry.items())[:5]:
            print(f"    {marker} → {text[:80]}...")

    # Show how many rows would be chunked
    last_row = (fn_start - 1) if fn_start else ws.max_row
    data_rows = last_row - data_start + 1
    print(f"\nChunkable data rows: {data_rows}")
```

if **name** == “**main**”:
import sys
import json

```
if len(sys.argv) < 2:
    print("Usage: python excel_chunker_v2.py <file.xlsx> [--debug]")
    sys.exit(1)

filepath = sys.argv[1]
debug_mode = "--debug" in sys.argv

if debug_mode:
    debug_sheet(filepath)
else:
    chunks = process_excel(filepath, inline_footnotes=True)
    print(f"\nTotal chunks: {len(chunks)}")
    for i, chunk in enumerate(chunks[:10]):  # show first 10
        print(f"\n--- Chunk {i+1} ---")
        print(f"Text: {chunk.text}")
        print(f"Metadata: {json.dumps(chunk.metadata, ensure_ascii=False, indent=2)}")
    if len(chunks) > 10:
        print(f"\n... and {len(chunks) - 10} more chunks")
```